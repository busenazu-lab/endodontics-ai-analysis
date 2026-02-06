import re
import math
import openpyxl
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

INPUT_XLSX = r"3. SINIF ÖDEVLERİ - Kod denemesi.xlsx"
OUTPUT_XLSX = r"3. SINIF ÖDEVLERİ - Kod denemesi_graded.xlsx"

MAX_COL = 20  # A..T

# ---------- Scoring (EDIT HERE if needed) ----------
LETTER_POINTS = {"A": 100, "B": 85, "C": 70, "D": 60, "E": 0}
PENALTY_PER_WEEK = 5
PASSING = 60

# Prototypes (RGB) for your late colors (nearest-match)
PROTOTYPES = {
    1: (209, 225, 211),  # green (1 week)
    2: (207, 236, 247),  # light blue (2 weeks)
    3: (246, 198, 173),  # light orange (3 weeks)
    4: (255, 0, 0),      # red (4 weeks)
    5: (117, 57, 25),    # brown (5 weeks)
    6: (166, 166, 166),  # gray (6 weeks)
    7: (64, 64, 64),     # blackish (7 weeks)
}

LETTER_RE = re.compile(r"[A-E]")

def dist2(a, b):
    return (a[0]-b[0])**2 + (a[1]-b[1])**2 + (a[2]-b[2])**2

def extract_letter(val):
    """Return A/B/C/D/E or None (blank/EKSİK)."""
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip().upper()
        if v == "" or "EKS" in v:
            return None
        m = LETTER_RE.search(v)
        return m.group(0) if m else None
    return None

def adjusted_score(letter, late_weeks):
    if letter is None:
        return 0
    base = LETTER_POINTS.get(letter, 0)
    return max(0, base - PENALTY_PER_WEEK * late_weeks)

# ---------- Theme color decoding (so Excel theme fills are readable) ----------
def get_theme_bases(wb):
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    root = ET.fromstring(wb.loaded_theme.decode("utf-8"))
    clr_scheme = root.find("a:themeElements/a:clrScheme", ns)
    order = ["lt1","dk1","lt2","dk2","accent1","accent2","accent3","accent4","accent5","accent6","hlink","folHlink"]
    bases = []
    for tag in order:
        el = clr_scheme.find(f"a:{tag}", ns)
        srgb = el.find(".//a:srgbClr", ns)
        if srgb is not None:
            rgb = srgb.attrib["val"]
        else:
            sys = el.find(".//a:sysClr", ns)
            rgb = sys.attrib.get("lastClr")
        bases.append(rgb)
    return bases

def apply_tint(rgb_hex, tint):
    r = int(rgb_hex[0:2], 16)
    g = int(rgb_hex[2:4], 16)
    b = int(rgb_hex[4:6], 16)

    def adj(c):
        if tint is None:
            return c
        if tint < 0:
            return int(round(c * (1.0 + tint)))
        return int(round(c * (1.0 - tint) + 255 * tint))

    r2 = max(0, min(255, adj(r)))
    g2 = max(0, min(255, adj(g)))
    b2 = max(0, min(255, adj(b)))
    return (r2, g2, b2)

def cell_fill_rgb(cell, theme_bases):
    fill = cell.fill
    if not fill or fill.patternType != "solid":
        return None

    fg = fill.fgColor
    if fg.type == "rgb" and fg.rgb and fg.rgb != "00000000":
        rgb = fg.rgb[-6:]
        return (int(rgb[0:2],16), int(rgb[2:4],16), int(rgb[4:6],16))

    if fg.type == "theme" and fg.theme is not None:
        base = theme_bases[fg.theme]
        return apply_tint(base, fg.tint)

    return None

def late_weeks_from_cell(cell, theme_bases):
    rgb = cell_fill_rgb(cell, theme_bases)
    if rgb is None:
        return 0
    week, _ = min(PROTOTYPES.items(), key=lambda kv: dist2(rgb, kv[1]))
    return week

# ---------- Workbook logic ----------
def get_assign_cols(ws):
    headers = [ws.cell(1, c).value for c in range(1, MAX_COL + 1)]

    name_col = None
    no_col = None
    for c, h in enumerate(headers, 1):
        if isinstance(h, str) and h.strip().lower() in ["adi soyadi", "adı soyadı", "ad soyad", "adi soyad"]:
            name_col = c
        if isinstance(h, str) and "öğrenci no" in h.lower():
            no_col = c

    if name_col is None:
        name_col = 1

    start = max(filter(None, [name_col, no_col])) + 1 if (no_col and name_col) else (name_col + 1)

    cols = []
    for c in range(start, MAX_COL + 1):
        h = headers[c - 1]
        # include if header exists or if there is any value in that column
        any_val = any(
            (ws.cell(r, c).value is not None and str(ws.cell(r, c).value).strip() != "")
            for r in range(2, ws.max_row + 1)
        )
        if not any_val and (h is None or str(h).strip() == ""):
            continue
        label = str(h).strip() if (h is not None and str(h).strip() != "") else get_column_letter(c)
        cols.append((c, label))

    return no_col, name_col, cols

def build_name_index(ws, name_col):
    idx = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if name is None:
            continue
        key = str(name).strip().upper()
        if key and key not in idx:
            idx[key] = r
    return idx

def main():
    wb = openpyxl.load_workbook(INPUT_XLSX)
    theme_bases = get_theme_bases(wb)

    sheets = ["ŞEKİLLENDİRME", "DOLUM", "Retreatment", "POST"]

    meta = {}
    for sh in sheets:
        ws = wb[sh]
        no_col, name_col, cols = get_assign_cols(ws)
        meta[sh] = {
            "ws": ws,
            "no_col": no_col,
            "name_col": name_col,
            "cols": cols,
            "index": build_name_index(ws, name_col),
        }

    # Roster from ŞEKİLLENDİRME (A,B)
    roster = []
    ws_base = wb["ŞEKİLLENDİRME"]
    for r in range(2, ws_base.max_row + 1):
        no = ws_base.cell(r, 1).value
        name = ws_base.cell(r, 2).value
        if name is None:
            continue
        roster.append((no, str(name).strip()))

    # Build summary rows
    rows = []
    for no, name in roster:
        key = name.upper()
        total = 0
        completed = 0
        all_scores = []
        missing_list = []

        per_sheet_avg = {}

        for sh in sheets:
            ws = meta[sh]["ws"]
            r = meta[sh]["index"].get(key)
            sheet_scores = []

            for c, label in meta[sh]["cols"]:
                total += 1
                if r is None:
                    missing_list.append(f"{sh}!{get_column_letter(c)} ({label})")
                    sheet_scores.append(0)
                    all_scores.append(0)
                    continue

                cell = ws.cell(r, c)
                letter = extract_letter(cell.value)
                if letter is None:
                    missing_list.append(f"{sh}!{get_column_letter(c)} ({label})")
                    sheet_scores.append(0)
                    all_scores.append(0)
                else:
                    completed += 1
                    lw = late_weeks_from_cell(cell, theme_bases)
                    sc = adjusted_score(letter, lw)
                    sheet_scores.append(sc)
                    all_scores.append(sc)

            per_sheet_avg[sh] = (sum(sheet_scores) / len(sheet_scores)) if sheet_scores else 0

        avg = (sum(all_scores) / len(all_scores)) if all_scores else 0
        status = "PASS" if avg >= PASSING else "FAIL"

        rows.append([
            no, name, total, completed, len(missing_list),
            ", ".join(missing_list),
            round(avg, 2), status,
            round(per_sheet_avg["ŞEKİLLENDİRME"], 2),
            round(per_sheet_avg["DOLUM"], 2),
            round(per_sheet_avg["Retreatment"], 2),
            round(per_sheet_avg["POST"], 2),
        ])

    # Write SUMMARY sheet
    if "SUMMARY" in wb.sheetnames:
        del wb["SUMMARY"]
    ws_sum = wb.create_sheet("SUMMARY")

    headers = [
        "Student No", "Name", "Total HW", "Completed", "Missing", "Missing Details",
        "Avg Score", "Status", "ŞEKİLLENDİRME Avg", "DOLUM Avg", "Retreatment Avg", "POST Avg"
    ]
    ws_sum.append(headers)
    for r in rows:
        ws_sum.append(r)

    # Light formatting
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = [12, 24, 10, 10, 10, 60, 10, 10, 18, 12, 16, 10]
    for c in range(1, len(headers) + 1):
        cell = ws_sum.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws_sum.column_dimensions[get_column_letter(c)].width = widths[c - 1]

    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")

    for r in range(2, ws_sum.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws_sum.cell(r, c)
            cell.border = border
            cell.alignment = left if c in [2, 6] else center

        st = ws_sum.cell(r, 8)
        st.fill = pass_fill if st.value == "PASS" else fail_fill
        sc = ws_sum.cell(r, 7)
        if isinstance(sc.value, (int, float)) and sc.value < PASSING:
            sc.fill = fail_fill

    ws_sum.freeze_panes = "A2"

    wb.save(OUTPUT_XLSX)
    print("Saved:", OUTPUT_XLSX)

if __name__ == "__main__":
    main()
